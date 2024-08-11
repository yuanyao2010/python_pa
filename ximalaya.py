
import requests
import json
import pprint
from openpyxl import load_workbook

headers = {'user-agent':
'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36'}

def get_data(page):
    url=f'https://www.ximalaya.com/revision/album/v1/getTracksList?albumId=11549955&pageNum={page}&pageSize=30'
    r = requests.get(url,headers=headers)
   # print(r.json())
   
    shu_data = r.json()
    data =  parse_data(shu_data)
    save_data(data)
    
    


def parse_data(data):
    track_data = []
    track_list = data['data']['tracks']
    for track in track_list:
        track_title = track['title']
        track_title = track_title.replace('丨','#') 
        track_title = track_title.replace('|', '#') 
        track_title = track_title.replace('｜', '#') 
        track_title_s = track_title.split("#",1)
        #track_title_s = track_title.split("|",1)
       # track_title_s =track_title.split("丨",1)
        track_data.append(track_title_s)
    return track_data
 

def save_data(data):
    wb = load_workbook(f'{file_path}/ximalaya.xlsx')
    ws = wb['Sheet1']
    for track in data:
        ws.append(track)
        print('save')
    wb.save(f'{file_path}/ximalaya.xlsx')
 
if __name__ == '__main__':
    file_path='D:/python_project'
    template_name='ximalaya.xlsx'
    wb = load_workbook(f'{file_path}/{template_name}')
    wb.save(f'{file_path}/ximalaya.xlsx')

    
    for page in range(75):
        get_data(page)

    print("end")
